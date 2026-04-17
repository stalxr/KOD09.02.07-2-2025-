import os
import openpyxl
from django.core.management.base import BaseCommand
from main.models import PartnerType, Partner, ProductType, Product, MaterialType, PartnerProduct
from decimal import Decimal


class Command(BaseCommand):
    help = 'Import data from Excel files'

    def handle(self, *args, **options):
        base_dir = r'c:\Users\_v.cheremnykh\Desktop\КОД 09.02.07-2-2025 Приложения к образцу задания Том 1\Ресурсы'

        self.import_partner_types(os.path.join(base_dir, 'Partners_import.xlsx'))
        self.import_product_types(os.path.join(base_dir, 'Product_type_import.xlsx'))
        self.import_material_types(os.path.join(base_dir, 'Material_type_import.xlsx'))
        self.import_products(os.path.join(base_dir, 'Products_import.xlsx'))
        self.import_partners(os.path.join(base_dir, 'Partners_import.xlsx'))
        self.import_partner_products(os.path.join(base_dir, 'Partner_products_import.xlsx'))

        self.stdout.write(self.style.SUCCESS('All data imported successfully!'))

    def import_partner_types(self, filepath):
        self.stdout.write('Importing partner types...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        types = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                types.add(row[0])

        for pt_name in types:
            PartnerType.objects.get_or_create(name=pt_name)

        self.stdout.write(self.style.SUCCESS(f'Created {len(types)} partner types'))

    def import_product_types(self, filepath):
        self.stdout.write('Importing product types...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ProductType.objects.get_or_create(
                    name=row[0],
                    defaults={'coefficient': row[1] or 0}
                )

        self.stdout.write(self.style.SUCCESS('Product types imported'))

    def import_material_types(self, filepath):
        self.stdout.write('Importing material types...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                MaterialType.objects.get_or_create(
                    name=row[0],
                    defaults={'defect_percentage': row[1] or 0}
                )

        self.stdout.write(self.style.SUCCESS('Material types imported'))

    def import_products(self, filepath):
        self.stdout.write('Importing products...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:
                product_type, _ = ProductType.objects.get_or_create(name=row[0])
                Product.objects.get_or_create(
                    article=str(row[2]),
                    defaults={
                        'name': row[1],
                        'product_type': product_type,
                        'min_price': row[3] or 0
                    }
                )

        self.stdout.write(self.style.SUCCESS('Products imported'))

    def import_partners(self, filepath):
        self.stdout.write('Importing partners...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:
                partner_type, _ = PartnerType.objects.get_or_create(name=row[0])
                Partner.objects.get_or_create(
                    inn=str(row[6]),
                    defaults={
                        'partner_type': partner_type,
                        'name': row[1],
                        'director': row[2],
                        'email': row[3],
                        'phone': row[4],
                        'address': row[5],
                        'rating': row[7] or 0
                    }
                )

        self.stdout.write(self.style.SUCCESS('Partners imported'))

    def import_partner_products(self, filepath):
        self.stdout.write('Importing partner product sales...')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                try:
                    partner = Partner.objects.get(name=row[1])
                    # Найдем продукт по имени
                    product = Product.objects.filter(name=row[0]).first()
                    if product and row[2]:
                        PartnerProduct.objects.create(
                            partner=partner,
                            product=product,
                            quantity=int(row[2]),
                            sale_date=row[3]
                        )
                except Partner.DoesNotExist:
                    continue

        self.stdout.write(self.style.SUCCESS('Partner product sales imported'))
